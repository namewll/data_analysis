from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os

font=Font(name='宋体',size=12,bold=True)
alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
side=Side(style="thin",color="000000")
border=Border(top=side,bottom=side,left=side,right=side)

if not os.path.exists("工作人员收入"):
    os.makedirs("工作人员收入")

wb = load_workbook('工资明细.xlsx')
ws = wb['Sheet']
index=0
title=[]

def write_data(title,res):
    wb=Workbook()
    ws=wb.active
    ws.append(title)
    ws.append(res)
    zimu="ABCDEFGHIJKL"
    for i in range(1,3):
        ws.row_dimensions[i].height = 30
        for j in zimu:
            ws.column_dimensions[j].width=11.22
            ws[f"{j}{i}"].font=font
            ws[f"{j}{i}"].alignment=alignment
            ws[f"{j}{i}"].border=border

    wb.save(f"工作人员收入/{res[1]}.xlsx")
    print(f"{res[1]}保存完成")

for row in ws:
    index+=1
    if index<3:
        continue
    if index==3:
        for cell in row:
            title.append(cell.value)
    res=[]
    if index>3:
        for cell in row:
            res.append(cell.value)
        write_data(title,res)




